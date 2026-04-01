#!/usr/bin/env python3
"""
Generador de TXT para Pago de Haberes - Banco Galicia
Replica la lógica del macro VBA "Generar TXT" del archivo .xlsm
"""

import sys
from datetime import datetime
import openpyxl

# --- Tablas de mapeo (hoja Validadores) ---

CONVENIO_CODE = {
    'Pago de Haberes': '*H3',
    'Pago a Proveedores': '*H3',
}

TIPO_CUENTA = {
    'Caja de Ahorro': 'A',
    'Cuenta Corriente': 'C',
    'Cuenta Cese Laboral': 'A',
}

MONEDA = {
    'Pesos': '1',
    'Dólares': '2',
}

CONSOLIDADO = {
    'Consolidado': 'S',
    'No Consolidado': 'N',
}

CONCEPTO = {
    'Acreditamiento Haberes': '01',
    'Horas Extra': '02',
    'Reintegro por Viáticos': '03',
    'Sueldo Anual Complementario': '04',
    'Subsidio Vacacional': '05',
    'Gastos de Representacion': '06',
    'Honorarios de Profesionales': '07',
    'Asignacion Personal Contratado': '08',
    'Asignacion Becas/Pasantias': '09',
    'Premio por Productividad/Calidad': '10',
    'Reembolso Gastos': '11',
    'Indemnizacion/Liquidacion Final': '12',
}

# --- Funciones de formato (equivalente al VBA CompletarLongitud) ---

def val_str(value):
    """Convierte cualquier valor de celda a string limpio."""
    if value is None:
        return ''
    if isinstance(value, float):
        return str(int(value))
    if isinstance(value, datetime):
        return value.strftime('%Y%m%d')
    return str(value).strip()

def pad_right(value, width):
    """Completar con espacios a la derecha (campo tipo 'Espacios')."""
    s = val_str(value)
    return s[:width].ljust(width)

def pad_left_zeros(value, width):
    """Completar con ceros a la izquierda (campo tipo 'Ceros a Izq')."""
    s = val_str(value)
    return s.zfill(width)[:width]  # trunca si el valor es más largo

def amount_str(value, width=14):
    """
    Convierte importe a string de enteros (sin punto decimal, 2 decimales implícitos).
    Ej: 1034549.33 → '00000103454933'
    """
    if not value:
        return '0' * width
    cents = round(float(value) * 100)
    return str(cents).zfill(width)

def fecha_str(value):
    """Convierte fecha Excel (datetime) a string AAAAMMDD."""
    if isinstance(value, datetime):
        return value.strftime('%Y%m%d')
    return val_str(value)

def cbu_str(value, width=26):
    """
    Convierte CBU (22 dígitos) a campo de 26 chars, ceros a la izquierda.
    Ej: '0720159888000002181086' → '0000720159888000002181086'
    """
    if value is None or value == '':
        return '0' * width
    digits = ''.join(c for c in str(value).strip() if c.isdigit())
    return digits.zfill(width)

# --- Generación de registros ---

def make_header(ws, fecha_override=None, importe_override=None):
    """
    Genera la línea de Header (477 chars).
    Fuente: hoja 'Def-Header' + celdas B2:B11 de 'Transferencias y Pagos'
    Si fecha_override está presente, se usa en lugar de B11.
    Si importe_override está presente, se usa en lugar de B10.
    """
    tipo_conv = val_str(ws['B2'].value)
    convenio  = pad_left_zeros(ws['B3'].value, 6)         # B3: N° Convenio
    cuit      = pad_left_zeros(ws['B4'].value, 11)         # B4: CUIT empresa (sin guiones)
    tipo_cta  = TIPO_CUENTA.get(val_str(ws['B5'].value), 'C')   # B5: Tipo Cuenta
    moneda    = MONEDA.get(val_str(ws['B6'].value), '1')         # B6: Moneda
    cta_deb   = pad_left_zeros(ws['B7'].value, 12)         # B7: Cuenta Débito
    leyenda   = pad_right(val_str(ws['B8'].value).upper(), 15)   # B8: Leyenda (upper)
    consol    = CONSOLIDADO.get(val_str(ws['B9'].value), 'N')    # B9: Consolidado
    importe   = importe_override if importe_override else amount_str(ws['B10'].value, 14)
    fecha     = fecha_override if fecha_override else fecha_str(ws['B11'].value)

    conv_code = CONVENIO_CODE.get(tipo_conv, '*H3')

    line = (
        conv_code   +   # pos   1- 3 (3)  Tipo de Convenio
        convenio    +   # pos   4- 9 (6)  Código de Empresa
        cuit        +   # pos  10-20 (11) CUIT
        tipo_cta    +   # pos  21   (1)   Tipo Cuenta Débito
        moneda      +   # pos  22   (1)   Moneda Débito
        cta_deb     +   # pos  23-34 (12) Cuenta Débito
        '0' * 26    +   # pos  35-60 (26) CBU Débito (no aplica → ceros)
        importe     +   # pos  61-74 (14) Importe Total
        fecha       +   # pos  75-82 (8)  Fecha Acreditación
        leyenda     +   # pos  83-97 (15) Leyenda
        consol      +   # pos  98   (1)   Débito Consolidado
        ' ' * 379       # pos  99-477 (379) Filler
    )
    assert len(line) == 477, f"Header length error: {len(line)}"
    return line


def make_detail(cols):
    """
    Genera la línea de Detalle (477 chars) para un empleado.
    cols: tupla con valores de columnas A-P de la fila del empleado.
    Fuente: hoja 'Def-Detalle' + '2-Detalle'
    """
    nombre      = cols[0]   # A: Nombre
    cuit_emp    = cols[1]   # B: CUIT
    fecha_acred = cols[2]   # C: Fecha Acreditación
    tipo_cta_c  = cols[3]   # D: Tipo Cuenta Crédito
    moneda_c    = cols[4]   # E: Moneda Crédito
    cuenta_cred = cols[5]   # F: Cuenta Crédito (vacío si usa CBU)
    cbu_cred    = cols[6]   # G: CBU Crédito (22 dígitos)
    importe     = cols[7]   # H: Importe
    leyenda_cred = cols[8]  # I: Leyenda Crédito
    id_interna  = cols[9]   # J: Identificación Interna
    fecha_proc  = cols[10]  # K: Fecha de Proceso
    concepto    = cols[11]  # L: Código de Concepto
    pago_com    = cols[12]  # M: Pago a Comercio
    nro_vep     = cols[13]  # N: Nro VEP
    leyenda_deb = cols[14]  # O: Leyenda Débito
    periodo     = cols[15]  # P: Periodo Cese Laboral

    nombre_s    = pad_right(nombre, 16)
    cuit_s      = pad_left_zeros(cuit_emp, 11)
    fec_acred_s = fecha_str(fecha_acred)
    tipo_cta_s  = TIPO_CUENTA.get(val_str(tipo_cta_c), 'A')
    moneda_s    = MONEDA.get(val_str(moneda_c), '1')
    cuenta_s    = cbu_str(cuenta_cred, 12)              # 12 zeros si vacío
    cbu_s       = cbu_str(cbu_cred, 26)                  # 26 chars
    importe_s   = amount_str(importe, 14)
    ley_cred_s  = pad_right(leyenda_cred, 15)
    id_int_s    = pad_right(id_interna, 22)
    fec_proc_s  = fecha_str(fecha_proc)
    concepto_s  = CONCEPTO.get(val_str(concepto), '01').zfill(2)
    pago_com_s  = pad_right(pago_com, 2)
    nro_vep_s   = pad_right(nro_vep, 14)
    email_s     = ' ' * 60
    ley_deb_s   = pad_right(val_str(leyenda_deb).upper(), 35)
    periodo_s   = pad_right(periodo, 15)

    line = (
        nombre_s    +   # pos   1-16  (16) Nombre
        cuit_s      +   # pos  17-27  (11) CUIT
        fec_acred_s +   # pos  28-35  (8)  Fecha Acreditación
        tipo_cta_s  +   # pos  36     (1)  Tipo Cuenta Crédito
        moneda_s    +   # pos  37     (1)  Moneda Crédito
        cuenta_s    +   # pos  38-49  (12) Cuenta Crédito
        cbu_s       +   # pos  50-75  (26) CBU Crédito
        '32'        +   # pos  76-77  (2)  Código Transacción (fijo)
        '1'         +   # pos  78     (1)  Tipo de Pago (fijo)
        importe_s   +   # pos  79-92  (14) Importe
        ley_cred_s  +   # pos  93-107 (15) Leyenda Crédito
        id_int_s    +   # pos 108-129 (22) Identificación Interna
        fec_proc_s  +   # pos 130-137 (8)  Fecha de Proceso
        concepto_s  +   # pos 138-139 (2)  Código de Concepto
        pago_com_s  +   # pos 140-141 (2)  Pago a Comercio
        nro_vep_s   +   # pos 142-155 (14) Nro VEP
        email_s     +   # pos 156-215 (60) Email Beneficiario
        ley_deb_s   +   # pos 216-250 (35) Leyenda Débito
        periodo_s   +   # pos 251-265 (15) Periodo Cese Laboral
        ' ' * 212       # pos 266-477 (212) Filler
    )
    assert len(line) == 477, f"Detail length error: {len(line)}"
    return line


def make_trailer(convenio, count):
    """
    Genera la línea de Trailer (477 chars).
    Fuente: hoja 'Def-Trailer' + '4-Trailer'
    """
    line = (
        '*F'                    +   # pos  1-2  (2)  Tipo de Registro
        str(convenio).zfill(6)  +   # pos  3-8  (6)  N° Convenio
        str(count).zfill(7)     +   # pos  9-15 (7)  Cant. Total Registros
        ' ' * 462                   # pos 16-477 (462) Filler
    )
    assert len(line) == 477, f"Trailer length error: {len(line)}"
    return line


# --- Main ---

def main(xlsm_path):
    print(f"Leyendo: {xlsm_path}")
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True, data_only=True)
    ws = wb['Transferencias y Pagos']

    # Datos empresa
    tipo_conv = val_str(ws['B2'].value)
    convenio  = val_str(ws['B3'].value)
    fecha_b11 = ws['B11'].value
    if isinstance(fecha_b11, datetime):
        fecha_larga = fecha_b11.strftime('%d%m%Y')   # DDMMYYYY para nombre de archivo
    else:
        fecha_larga = val_str(fecha_b11)

    # Empleados (filas 15 en adelante, col A con nombre — incluye importe 0, igual que el VBA)
    employees = []
    for row in range(15, 10000):
        nombre = ws.cell(row=row, column=1).value
        if nombre is None:
            break
        cols = tuple(ws.cell(row=row, column=c).value for c in range(1, 17))
        employees.append(cols)

    # Validar que la fecha del header (B11) coincida con la fecha de detalle (col C).
    # Si no coinciden, usar la fecha del primer empleado y avisar.
    fecha_header = fecha_str(fecha_b11)
    if employees:
        fecha_detalle = fecha_str(employees[0][2])  # col C del primer empleado
        if fecha_header != fecha_detalle:
            print(f"  AVISO: Fecha Header B11 ({fecha_header}) no coincide con "
                  f"fecha detalle ({fecha_detalle}). Usando fecha detalle.")
            fecha_header = fecha_detalle
            if isinstance(employees[0][2], datetime):
                fecha_larga = employees[0][2].strftime('%d%m%Y')

    # Líneas de detalle
    details = [make_detail(emp) for emp in employees]

    # Validar que el importe del header (B10) coincida con la suma de importes
    # del detalle. Si no coinciden, usar la suma real para evitar rechazos bancarios.
    importe_header = amount_str(ws['B10'].value, 14)
    suma_detalle = sum(round(float(emp[7] or 0) * 100) for emp in employees)
    importe_calculado = str(suma_detalle).zfill(14)
    importe_override = None
    if importe_header != importe_calculado:
        diff_cents = abs(int(importe_header) - suma_detalle)
        print(f"  AVISO: Importe Header B10 ({importe_header}) no coincide con "
              f"suma detalle ({importe_calculado}). Diff: ${diff_cents/100:.2f}. "
              f"Usando suma detalle.")
        importe_override = importe_calculado

    # Header (usa fecha e importe validados contra detalle)
    header = make_header(ws, fecha_override=fecha_header, importe_override=importe_override)

    # Trailer
    trailer = make_trailer(convenio, len(details))

    # Nombre del archivo de salida
    output_name = f"{tipo_conv}_{convenio}_{fecha_larga}.txt"

    # Escribir archivo: UTF-8 sin BOM, cada línea termina en CRLF
    # (Banco Galicia no acepta BOM — el archivo de feb que funcionó no lo tiene)
    with open(output_name, 'wb') as f:
        for line in [header] + details + [trailer]:
            f.write(line.encode('utf-8') + b'\r\n')

    print(f"Archivo generado: {output_name}")
    print(f"  Header   : 1 registro")
    print(f"  Detalle  : {len(details)} empleados")
    print(f"  Trailer  : 1 registro")
    print(f"  Total pesos : ${suma_detalle/100:,.2f}")


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print(f"Uso: python3 {sys.argv[0]} archivo.xlsm")
        sys.exit(1)
    main(sys.argv[1])
